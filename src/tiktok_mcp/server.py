"""
Real TikTok for Business MCP server — our own code, not a vendored third-party
repo (see README.md for why). Wraps TikTok's documented Marketing API v1.3
(business-api.tiktok.com) as MCP tools, run as a local stdio subprocess.

Auth: standard OAuth2 authorization-code flow against a TikTok Developer app
(App ID + Secret) that the owner creates themselves at
business-api.tiktok.com/portal — that step needs a human login + consent
click and can't be automated.

    1. tiktok_ads_login()                    -> authorize URL to open in a browser
    2. (owner approves, copies the `auth_code` query param off the redirect URL)
    3. tiktok_ads_complete_auth(auth_code)    -> exchanges it, persists the
       access token + advertiser id to .env so every future process picks it
       up (same persistence pattern as src/mcp_tools/shopify_auth.py's
       persist_shopify_token)

Read-only reporting once authenticated — this agent never creates, edits, or
pauses a campaign:
    - list_campaigns()
    - get_ads_report(start_date, end_date) -> spend/impressions/clicks/CTR/
      conversions per TikTok's integrated report endpoint. Never fabricates a
      metric TikTok didn't actually return (e.g. ROAS needs a purchase-value
      pixel/catalog wired up on TikTok's side — omitted if absent).

WHILE THE APP IS STILL PENDING APPROVAL there is no API path at all: every
endpoint above requires an `Access-Token` header that only comes out of the
OAuth exchange, and the Ads Manager web session is not a substitute (curl'ing
it with browser cookies hits undocumented internal endpoints that change
without notice). The supported bridge in the meantime is the CSV/XLSX the owner
exports from Ads Manager by hand:

    - read_ads_export() -> parses the newest export dropped in
      data/tiktok_exports/ (override with TIKTOK_EXPORT_DIR) into the same
      metric shape get_ads_report returns, so the agent reports REAL numbers
      from the owner's own account with zero approval. Totals are summed from
      the raw columns and ratios (CTR/CPC/CPA/ROAS) are derived here rather
      than trusted from the file, since Ads Manager's own ratio columns are
      per-row averages that don't sum.
"""
from __future__ import annotations

import csv
import json
import os
import re
from pathlib import Path

import httpx
from mcp.server.fastmcp import FastMCP

_API_BASE = "https://business-api.tiktok.com/open_api/v1.3"
_ENV_PATH = Path(__file__).resolve().parents[2] / ".env"

mcp = FastMCP("tiktok-ads")


def _app_id() -> str:
    return os.environ.get("TIKTOK_APP_ID", "").strip()


def _app_secret() -> str:
    return os.environ.get("TIKTOK_APP_SECRET", "").strip()


def _redirect_uri() -> str:
    # TikTok's OAuth app rejects localhost/loopback redirect URIs outright
    # (confirmed live in the Developer Portal) — must be a real domain, even
    # though nothing needs to actually be listening there (see README: the
    # flow is copying the auth_code straight off the browser's address bar).
    return os.environ.get("TIKTOK_OAUTH_REDIRECT", "https://alphaforbaby.com/tiktok/callback").strip()


def _access_token() -> str:
    return os.environ.get("TIKTOK_ACCESS_TOKEN", "").strip()


def _advertiser_id() -> str:
    return os.environ.get("TIKTOK_ADVERTISER_ID", "").strip()


def _persist(**kv: str) -> None:
    """Write/replace KEY=value lines in .env and refresh this process's env so a
    value minted this call is usable immediately (mirrors shopify_auth's
    persist_shopify_token — same "one source of truth" convention)."""
    lines = _ENV_PATH.read_text().splitlines() if _ENV_PATH.exists() else []
    for key, value in kv.items():
        prefix = f"{key}="
        for i, ln in enumerate(lines):
            if ln.startswith(prefix):
                lines[i] = f"{prefix}{value}"
                break
        else:
            lines.append(f"{prefix}{value}")
        os.environ[key] = value
    _ENV_PATH.write_text("\n".join(lines) + "\n")


def _headers() -> dict:
    return {"Access-Token": _access_token(), "Content-Type": "application/json"}


@mcp.tool()
def tiktok_ads_login() -> dict:
    """Start OAuth: returns the URL to open in a browser and approve access.
    Requires TIKTOK_APP_ID to already be set — create a TikTok Developer app
    at business-api.tiktok.com/portal, add the Marketing API product, and
    request read-only Reporting + Ad Account scopes."""
    app_id = _app_id()
    if not app_id:
        return {
            "error": "TIKTOK_APP_ID not set — create a TikTok Developer app first at "
                     "business-api.tiktok.com/portal, then set TIKTOK_APP_ID/TIKTOK_APP_SECRET in .env"
        }
    url = (
        "https://business-api.tiktok.com/portal/auth"
        f"?app_id={app_id}&state=alpha-shoop&redirect_uri={_redirect_uri()}"
    )
    return {
        "authorize_url": url,
        "next_step": "Open this URL, log in, approve access, then copy the `auth_code` "
                     "query param from the redirect URL and call tiktok_ads_complete_auth with it.",
    }


@mcp.tool()
async def tiktok_ads_complete_auth(auth_code: str) -> dict:
    """Finish OAuth: exchange the `auth_code` from the redirect URL for an
    access token, and persist it (+ the advertiser id it's scoped to) to .env
    so every future call is already authenticated."""
    app_id, secret = _app_id(), _app_secret()
    if not app_id or not secret:
        return {"error": "TIKTOK_APP_ID / TIKTOK_APP_SECRET not set"}
    async with httpx.AsyncClient(timeout=20) as c:
        r = await c.post(
            f"{_API_BASE}/oauth2/access_token/",
            json={
                "app_id": app_id, "secret": secret,
                "auth_code": auth_code, "grant_type": "authorization_code",
            },
        )
    body = r.json()
    if body.get("code") != 0:
        return {"error": f"TikTok token exchange failed: {body.get('message')}", "raw": body}
    data = body["data"]
    token = data["access_token"]
    advertiser_ids = data.get("advertiser_ids", [])
    _persist(TIKTOK_ACCESS_TOKEN=token)
    if advertiser_ids and not _advertiser_id():
        _persist(TIKTOK_ADVERTISER_ID=str(advertiser_ids[0]))
    return {"authenticated": True, "advertiser_ids": advertiser_ids}


@mcp.tool()
def tiktok_ads_auth_status() -> dict:
    """Whether TikTok Ads is currently authenticated (has a stored access token)."""
    return {"authenticated": bool(_access_token()), "advertiser_id": _advertiser_id() or None}


@mcp.tool()
async def list_campaigns(advertiser_id: str = "") -> dict:
    """List TikTok ad campaigns (id, name, status, objective, budget) for the
    advertiser account. Read-only. Requires a completed tiktok_ads_login +
    tiktok_ads_complete_auth first."""
    token = _access_token()
    adv = advertiser_id or _advertiser_id()
    if not token or not adv:
        return {"error": "Not authenticated — run tiktok_ads_login first"}
    async with httpx.AsyncClient(timeout=20) as c:
        r = await c.get(
            f"{_API_BASE}/campaign/get/",
            params={"advertiser_id": adv, "page": 1, "page_size": 50},
            headers=_headers(),
        )
    body = r.json()
    if body.get("code") != 0:
        return {"error": f"TikTok campaign/get failed: {body.get('message')}", "raw": body}
    return {"campaigns": body["data"].get("list", [])}


@mcp.tool()
async def get_ads_report(start_date: str, end_date: str, advertiser_id: str = "") -> dict:
    """Real spend / impressions / clicks / CTR / conversions for the account
    between start_date and end_date (YYYY-MM-DD), via TikTok's integrated
    report endpoint. Read-only. Requires a completed tiktok_ads_login +
    tiktok_ads_complete_auth first. Never fabricates a metric TikTok didn't
    actually return."""
    token = _access_token()
    adv = advertiser_id or _advertiser_id()
    if not token or not adv:
        return {"error": "Not authenticated — run tiktok_ads_login first"}
    metrics = ["spend", "impressions", "clicks", "ctr", "conversion", "cost_per_conversion"]
    async with httpx.AsyncClient(timeout=20) as c:
        r = await c.get(
            f"{_API_BASE}/report/integrated/get/",
            params={
                "advertiser_id": adv,
                "report_type": "BASIC",
                "data_level": "AUCTION_ADVERTISER",
                "dimensions": json.dumps(["advertiser_id"]),
                "metrics": json.dumps(metrics),
                "start_date": start_date,
                "end_date": end_date,
                "page": 1,
                "page_size": 10,
            },
            headers=_headers(),
        )
    body = r.json()
    if body.get("code") != 0:
        return {"error": f"TikTok report/integrated/get failed: {body.get('message')}", "raw": body}
    rows = body["data"].get("list", [])
    return {"start_date": start_date, "end_date": end_date, "rows": rows}


# ── Manual CSV export bridge (works with no API approval) ─────────────────────

def _export_dir() -> Path:
    raw = os.environ.get("TIKTOK_EXPORT_DIR", "").strip()
    root = Path(raw) if raw else Path(__file__).resolve().parents[2] / "data" / "tiktok_exports"
    return root


# Raw additive columns only, matched IN ORDER — the first match wins, so the
# most specific patterns come first. Ratio/derived columns are deliberately not
# matched here (filtered by _is_derived) because summing or averaging them
# across rows is wrong; we recompute them from the totals instead.
#
# Order matters: "Total purchase value" contains "purchase", so a naive
# conversions-first match counts money as a conversion count (it did — a test
# export reported 6,094 conversions instead of 108, and lost ROAS entirely).
# Any money-shaped column ("... value") must be classified as revenue first.
_METRIC_ALIASES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("revenue", ("purchase value", "payment value", "conversion value", "revenue",
                 "value")),
    ("spend", ("cost", "spend", "amount spent")),
    ("impressions", ("impression", "show")),
    ("clicks", ("click",)),
    ("conversions", ("conversion", "complete payment", "purchase", "result")),
)

# Anything ratio-shaped: a per-unit cost, a rate, or a percentage.
_DERIVED_RE = re.compile(
    r"\bper\b|rate|ctr|cvr|cpc|cpm|cpa|roas|%|percent|average|avg", re.IGNORECASE
)


def _is_derived(header: str) -> bool:
    return bool(_DERIVED_RE.search(header))


# The column to group by, most granular first — comparing two ads needs a per-ad
# split, not one account-level total.
_DIMENSION_HEADERS = ("ad name", "creative name", "ad group name", "adgroup name", "campaign name")


def _dimension_index(headers: list[str]) -> tuple[int, str] | None:
    """(column index, label) of the most granular name column present."""
    lowered = [h.strip().lower() for h in headers]
    for needle in _DIMENSION_HEADERS:
        for i, h in enumerate(lowered):
            if h == needle or h.startswith(needle):
                return i, headers[i].strip()
    return None


def _derive(m: dict[str, float]) -> dict[str, float]:
    """Ratios computed from totals. The export's own ratio columns are per-row
    and cannot be summed, so they're ignored and recomputed here."""
    d: dict[str, float] = {}
    if m.get("impressions"):
        d["ctr_pct"] = round(m.get("clicks", 0) / m["impressions"] * 100, 3)
        d["cpm"] = round(m.get("spend", 0) / m["impressions"] * 1000, 2)
    if m.get("clicks"):
        d["cpc"] = round(m.get("spend", 0) / m["clicks"], 3)
    if m.get("conversions"):
        d["cpa"] = round(m.get("spend", 0) / m["conversions"], 2)
    # ROAS only when the export actually carries purchase value — never guessed.
    if m.get("revenue") and m.get("spend"):
        d["roas"] = round(m["revenue"] / m["spend"], 2)
    return d


def _rank_ads(groups: dict[str, dict[str, float]]) -> tuple[list[dict], str]:
    """Rank ads best-first and say which metric decided it.

    Picks the most meaningful metric the data actually supports: ROAS when
    purchase value was exported, else cost-per-acquisition, else CTR. Ads with
    no spend are excluded — a nil-spend row says nothing about performance."""
    rows = []
    for name, m in groups.items():
        d = _derive(m)
        rows.append({"ad": name, "metrics": {k: round(v, 2) for k, v in m.items()}, "derived": d})

    spending = [r for r in rows if r["metrics"].get("spend", 0) > 0] or rows
    if any("roas" in r["derived"] for r in spending):
        metric = "roas"
        spending.sort(key=lambda r: r["derived"].get("roas", -1), reverse=True)
    elif any("cpa" in r["derived"] for r in spending):
        metric = "cpa (lower is better)"
        spending.sort(key=lambda r: r["derived"].get("cpa", float("inf")))
    elif any("ctr_pct" in r["derived"] for r in spending):
        metric = "ctr (no conversion data in this export)"
        spending.sort(key=lambda r: r["derived"].get("ctr_pct", -1), reverse=True)
    else:
        metric = "spend only — this export has no performance columns"
    return spending, metric


def _classify(header: str) -> str:
    """Map one export column header to a canonical metric name ("" if none).

    Ads Manager column names differ by report type, account language and
    optimisation goal ("Cost" vs "Total cost", "Clicks (destination)",
    "Complete payment"), so match on substrings rather than exact names."""
    h = header.strip().lower()
    if not h or _is_derived(h):
        return ""
    for metric, needles in _METRIC_ALIASES:
        if any(n in h for n in needles):
            return metric
    return ""


# Ads Manager appends a grand-total row ("Total of 2 results"). Summing it along
# with the data rows doubles every metric, so it must be dropped — it is a
# summary of the rows above it, not another ad.
_TOTAL_ROW_RE = re.compile(r"^\s*total\b|^\s*grand total\b|^\s*总计", re.IGNORECASE)


def _is_total_row(row: list[str], dim_idx: int | None) -> bool:
    if dim_idx is not None and dim_idx < len(row):
        return bool(_TOTAL_ROW_RE.match(str(row[dim_idx])))
    return any(_TOTAL_ROW_RE.match(str(c)) for c in row[:2])


def _num(value: str) -> float | None:
    """Parse an export cell into a number: strips currency symbols, thousands
    separators and stray whitespace. Returns None for blanks/'--'/text."""
    if value is None:
        return None
    s = str(value).strip().replace(",", "").replace(" ", "")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in ("-", ".", "--"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _read_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    """Return (headers, rows) for a .csv or .xlsx export."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # noqa: BLE001
            raise RuntimeError(
                f"{path.name} is a spreadsheet and openpyxl isn't installed — "
                "re-export from Ads Manager as CSV, or `pip install openpyxl`"
            ) from exc
        # NOT read_only: TikTok's exports omit the sheet dimension record, and
        # openpyxl's read-only worksheet then yields no rows at all (silently —
        # it reported a real 2-ad report as "empty"). The files are tiny, so
        # loading them fully costs nothing.
        ws = load_workbook(path, data_only=True).active
        rows = [["" if c is None else str(c) for c in r] for r in ws.iter_rows(values_only=True)]
        return (rows[0], rows[1:]) if rows else ([], [])
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    reader = list(csv.reader(text.splitlines()))
    return (reader[0], reader[1:]) if reader else ([], [])


@mcp.tool()
def read_ads_export(filename: str = "") -> dict:
    """Read REAL TikTok Ads numbers from a report exported by hand out of Ads
    Manager — the only path to real data while the Developer app is still
    pending approval. Drop the export in data/tiktok_exports/ (or set
    TIKTOK_EXPORT_DIR); with no `filename` the newest file is used. Sums the
    raw columns and derives CTR/CPC/CPA/ROAS from those totals. Never invents a
    metric the file doesn't contain."""
    directory = _export_dir()
    if not directory.exists():
        return {"error": f"No export folder yet — create {directory} and drop the "
                         "CSV you export from TikTok Ads Manager into it."}
    if filename:
        path = directory / filename
        if not path.exists():
            return {"error": f"{filename} not found in {directory}"}
    else:
        files = [p for p in directory.iterdir()
                 if p.is_file() and p.suffix.lower() in (".csv", ".xlsx", ".xls")]
        if not files:
            return {"error": f"No CSV/XLSX export found in {directory} — export a report "
                             "from TikTok Ads Manager and drop it in there."}
        path = max(files, key=lambda p: p.stat().st_mtime)

    try:
        headers, rows = _read_rows(path)
    except Exception as exc:  # noqa: BLE001
        return {"error": f"Couldn't read {path.name}: {exc}"}
    if not headers or not rows:
        return {"error": f"{path.name} is empty"}

    columns = {i: _classify(h) for i, h in enumerate(headers)}
    matched = {m for m in columns.values() if m}
    if not matched:
        return {"error": f"{path.name} has no recognisable metric columns "
                         f"(saw: {', '.join(headers[:12])})"}

    date_idx = next(
        (i for i, h in enumerate(headers) if h.strip().lower() in ("date", "day", "time", "date range")),
        None,
    )
    dim = _dimension_index(headers)
    totals: dict[str, float] = {m: 0.0 for m in matched}
    per_ad: dict[str, dict[str, float]] = {}
    dates: list[str] = []
    counted = 0
    skipped_totals = 0
    for row in rows:
        if not any(str(c).strip() for c in row):
            continue
        if _is_total_row(row, dim[0] if dim else None):
            skipped_totals += 1
            continue
        counted += 1
        bucket = None
        if dim is not None and dim[0] < len(row):
            label = str(row[dim[0]]).strip()
            if label:
                bucket = per_ad.setdefault(label, {m: 0.0 for m in matched})
        for i, metric in columns.items():
            if not metric or i >= len(row):
                continue
            v = _num(row[i])
            if v is not None:
                totals[metric] += v
                if bucket is not None:
                    bucket[metric] += v
        if date_idx is not None and date_idx < len(row) and str(row[date_idx]).strip():
            dates.append(str(row[date_idx]).strip())

    out: dict[str, object] = {
        "source": "manual_export",
        "file": path.name,
        "rows_counted": counted,
        "total_rows_skipped": skipped_totals,
        "metrics": {k: round(v, 2) for k, v in totals.items()},
    }
    if dates:
        out["start_date"], out["end_date"] = min(dates), max(dates)
    out["derived"] = _derive(totals)

    # Per-ad breakdown + which one actually won. This is the question that
    # matters when deciding what to upload more of — an account-level total
    # can't distinguish a winning creative from one that's burning budget.
    if per_ad:
        ranked, metric = _rank_ads(per_ad)
        out["grouped_by"] = dim[1] if dim else None
        out["ranking_metric"] = metric
        out["ads"] = ranked
        if ranked:
            out["best_ad"] = ranked[0]["ad"]
            if len(ranked) > 1:
                out["worst_ad"] = ranked[-1]["ad"]
    return out


@mcp.tool()
def ads_export_status() -> dict:
    """Where to drop manual Ads Manager exports, and what's there right now."""
    directory = _export_dir()
    files = []
    if directory.exists():
        files = sorted(
            (p.name for p in directory.iterdir()
             if p.is_file() and p.suffix.lower() in (".csv", ".xlsx", ".xls")),
        )
    return {"export_dir": str(directory), "exists": directory.exists(), "files": files}


if __name__ == "__main__":
    mcp.run()
